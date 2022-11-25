import requests, json
#from openpyxl #import Workbook
from openpyxl import load_workbook
from datetime import datetime
def tiempo_pronostico():

    print ("Script 4 para pronostico del tiempo")
    lat = "34.0430219"
    lon = "118.2694428"
    appid = "c2f7636c754671lo23412af60c4dc8deae" #Mi api
    page = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat="+lat+"&lon="+lon+"&units=metric&lang=sp&exclude=hourly,minutely,current&appid="+appid)
    if page.status_code != 200:
        print("Página no encontrada, fin del script")
    else:
        wheatherData = json.loads(page.content)
        print("Pronóstico para los días de los siguientes partidos de Los Angeles Lakers")
        print("*****Partido 1, Estadio STAPLES CENTER*****")
        print("Para el partido Lakers vs Suns, el día 26 de Mayo:")
        print("Temperatura del día: ",wheatherData['daily'][1]["temp"]["day"])
        print("Descripción del día en general:",wheatherData['daily'][1]["weather"][0]["description"])
        print("\n")
        print("*****Partido 2, Estadio Phoenix Suns Arena*****")
        print("Para el partido Lakers vs Suns, el día 1 de junio:")
        print("Temperatura del día: ",wheatherData['daily'][6]["temp"]["day"])
        print("Descripción del día en general:",wheatherData['daily'][6]["weather"][0]["description"])
        print("\n")
        print("*****Partido 3, Estadio STAPLES CENTER*****")
        print("Para el partido Lakers vs Broklyn Nets, el día 5 de junio:")
        print("Temperatura del día: ",wheatherData['daily'][7]["temp"]["day"])
        print("Descripción del día en general:",wheatherData['daily'][7]["weather"][0]["description"])
        clima = {"Partido 1": "Lakers",
                 "Día1": "8/12",
                 "Estadio1" : "STAPLES CENTER, Los Angeles California",
                 "Clima Partido 1": wheatherData['daily'][1]["temp"]["day"],
                 "Des. gen. del día 1": wheatherData['daily'][1]["weather"][0]["description"],
                 "Clima Partido 2": wheatherData['daily'][6]["temp"]["day"],
                 "Partido 2": "Suns",
                 "Día 2": "13/12",
                 "Estadio 2" : "Phoenix suns Arena, Arizona",
                 "Des. gen. del día 2": wheatherData['daily'][6]["weather"][0]["description"],
                 "Partido 3": "Nets",
                 "Día 3": "14/12",
                 "Estadio 3" : "STAPLES CENTER, Los Angeles california",
                 "Clima Partido 3": wheatherData['daily'][7]["temp"]["day"],
                 "Des. gen. del día 3":wheatherData['daily'][7]["weather"][0]["description"]}
        ws = load_workbook("EXCEL PIA IP.xlsx")
        hoja3 = ws.create_sheet("Clima",3)
        #hoja3 = libro.active
        #cadena = "Clima"
        #hoja3.title = cadena
        w2 = ws["Clima"]
        w2["B1"] = "Partido 1"
        w2["C1"] = "Partido 2"
        w2["D1"] = "Partido 3"
        w2["A2"] = "Día"
        w2["A3"] = "Lakers vs "
        w2["A4"] = "Estadio"
        w2["A5"] = "Clima Celsius"
        w2["A6"] = "Des. gen. del día"
        w2["C2"] = clima["Día 2"]
        w2["D2"] = clima["Día 3"]
        w2["B2"] = clima["Día1"]
        w2["B3"] = clima["Partido 1"]
        w2["B4"] = clima["Estadio1"]
        w2["B5"] = clima["Clima Partido 1"]
        w2["B6"] = clima["Des. gen. del día 1"]
        w2["C2"] = clima["Día 2"]
        w2["C3"] = clima["Partido 2"]
        w2["C4"] = clima["Estadio 2"]
        w2["C5"] = clima["Clima Partido 2"]
        w2["C6"] = clima["Des. gen. del día 2"]
        w2["D2"] = clima["Día 3"]
        w2["D3"] = clima["Partido 3"]
        w2["D4"] = clima["Estadio 3"]
        w2["D5"] = clima["Clima Partido 3"]
        w2["D6"] = clima["Des. gen. del día 3"]
        ws.save("EXCEL PIA IP.xlsx") 
        
tiempo_pronostico()
print("Script Concluido:) ")
    
    #entonces el 0 es el de hoy, a lo que el 7 en USA, 6 en mx, el 1 es el 8, y así sucesivamente
    
